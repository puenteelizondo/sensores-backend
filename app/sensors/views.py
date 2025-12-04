from rest_framework import viewsets, status
from rest_framework.response import Response
from rest_framework.decorators import action
from django.http import HttpResponse
from .models import Temperatura, Humedad, Distancia
from .serializers import TemperaturaSerializer, HumedadSerializer, DistanciaSerializer
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime

class TemperaturaViewSet(viewsets.ModelViewSet):
    queryset = Temperatura.objects.all()
    serializer_class = TemperaturaSerializer
    
    @action(detail=False, methods=['get'])
    def ultima(self, request):
        """Obtiene la última lectura de temperatura"""
        ultima_temp = self.queryset.order_by('-id').first()

        if ultima_temp:
            serializer = self.get_serializer(ultima_temp)
            return Response(serializer.data)
        return Response({'mensaje': 'No hay datos disponibles'}, status=status.HTTP_404_NOT_FOUND)

class HumedadViewSet(viewsets.ModelViewSet):
    queryset = Humedad.objects.all()
    serializer_class = HumedadSerializer
    
    @action(detail=False, methods=['get'])
    def ultima(self, request):
        """Obtiene la última lectura de humedad"""
        ultima_hum = self.queryset.order_by('-id').first()

        if ultima_hum:
            serializer = self.get_serializer(ultima_hum)
            return Response(serializer.data)
        return Response({'mensaje': 'No hay datos disponibles'}, status=status.HTTP_404_NOT_FOUND)

class DistanciaViewSet(viewsets.ModelViewSet):
    queryset = Distancia.objects.all()
    serializer_class = DistanciaSerializer
    
    @action(detail=False, methods=['get'])
    def ultima(self, request):
        """Obtiene la última lectura de distancia"""
        ultima_dist = self.queryset.order_by('-id').first()

        if ultima_dist:
            serializer = self.get_serializer(ultima_dist)
            return Response(serializer.data)
        return Response({'mensaje': 'No hay datos disponibles'}, status=status.HTTP_404_NOT_FOUND)


# ============================================
# NUEVO: ViewSet para reporte Excel
# ============================================
from rest_framework import viewsets
from rest_framework.decorators import action

class ReporteViewSet(viewsets.ViewSet):
    """
    ViewSet para generar reportes de sensores
    """
    
    @action(detail=False, methods=['get'])
    def excel(self, request):
        """
        Genera un reporte Excel con datos de todos los sensores
        GET /api/reporte/excel/
        
        Parámetros opcionales:
        - limite: número de registros por sensor (default: 100)
        """
        limite = int(request.query_params.get('limite', 100))
        
        # Obtener datos
        temperaturas = Temperatura.objects.all()[:limite]
        humedades = Humedad.objects.all()[:limite]
        distancias = Distancia.objects.all()[:limite]
        
        # Crear workbook
        wb = openpyxl.Workbook()
        
        # ===== HOJA 1: RESUMEN =====
        ws_resumen = wb.active
        ws_resumen.title = "📊 Resumen"
        
        # Estilos
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=12)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Título principal
        ws_resumen['A1'] = 'REPORTE DE SENSORES IoT'
        ws_resumen['A1'].font = Font(bold=True, size=16, color="1F4E78")
        ws_resumen['A1'].alignment = Alignment(horizontal='center')
        ws_resumen.merge_cells('A1:D1')
        
        ws_resumen['A2'] = f'Generado: {datetime.now().strftime("%Y-%m-%d %H:%M:%S")}'
        ws_resumen['A2'].font = Font(italic=True, size=10)
        ws_resumen.merge_cells('A2:D2')
        
        # Headers resumen
        headers_resumen = ['Sensor', 'Total Lecturas', 'Última Lectura', 'Promedio']
        for col, header in enumerate(headers_resumen, 1):
            cell = ws_resumen.cell(row=4, column=col)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')
            cell.border = border
        
        # Datos resumen
        datos_resumen = [
            ['🌡️ Temperatura', 
             temperaturas.count(), 
             f"{temperaturas.first().valor}°C" if temperaturas.exists() else "N/A",
             f"{sum(t.valor for t in temperaturas) / temperaturas.count():.2f}°C" if temperaturas.exists() else "N/A"],
            ['💧 Humedad', 
             humedades.count(), 
             f"{humedades.first().valor}%" if humedades.exists() else "N/A",
             f"{sum(h.valor for h in humedades) / humedades.count():.2f}%" if humedades.exists() else "N/A"],
            ['📏 Distancia', 
             distancias.count(), 
             f"{distancias.first().valor}cm" if distancias.exists() else "N/A",
             f"{sum(d.valor for d in distancias) / distancias.count():.2f}cm" if distancias.exists() else "N/A"],
        ]
        
        fills_resumen = [
            PatternFill(start_color="FFE699", end_color="FFE699", fill_type="solid"),  # Amarillo
            PatternFill(start_color="B4C7E7", end_color="B4C7E7", fill_type="solid"),  # Azul claro
            PatternFill(start_color="C6E0B4", end_color="C6E0B4", fill_type="solid"),  # Verde claro
        ]
        
        for row_idx, (datos, fill) in enumerate(zip(datos_resumen, fills_resumen), 5):
            for col_idx, valor in enumerate(datos, 1):
                cell = ws_resumen.cell(row=row_idx, column=col_idx)
                cell.value = valor
                cell.fill = fill
                cell.border = border
                cell.alignment = Alignment(horizontal='center')
        
        # Ajustar anchos
        for col in range(1, 5):
            ws_resumen.column_dimensions[get_column_letter(col)].width = 20
        
        # ===== HOJA 2: TEMPERATURA =====
        ws_temp = wb.create_sheet(title="🌡️ Temperatura")
        self._crear_hoja_sensor(
            ws_temp, 
            temperaturas, 
            "TEMPERATURA",
            PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid"),
            PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid"),
            "°C"
        )
        
        # ===== HOJA 3: HUMEDAD =====
        ws_hum = wb.create_sheet(title="💧 Humedad")
        self._crear_hoja_sensor(
            ws_hum, 
            humedades, 
            "HUMEDAD",
            PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid"),
            PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid"),
            "%"
        )
        
        # ===== HOJA 4: DISTANCIA =====
        ws_dist = wb.create_sheet(title="📏 Distancia")
        self._crear_hoja_sensor(
            ws_dist, 
            distancias, 
            "DISTANCIA",
            PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid"),
            PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid"),
            "cm"
        )
        
        # Guardar en memoria y retornar
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="reporte_sensores_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx"'
        wb.save(response)
        
        return response
    
    def _crear_hoja_sensor(self, ws, queryset, titulo, header_fill, row_fill, unidad):
        """Método auxiliar para crear hojas de sensores individuales"""
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Título
        ws['A1'] = titulo
        ws['A1'].font = Font(bold=True, size=14, color="1F4E78")
        ws['A1'].alignment = Alignment(horizontal='center')
        ws.merge_cells('A1:D1')
        
        # Headers
        headers = ['ID', 'Valor', 'Fecha y Hora', 'Sensor ID']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col)
            cell.value = header
            cell.fill = header_fill
            cell.font = Font(bold=True, color="FFFFFF")
            cell.alignment = Alignment(horizontal='center')
            cell.border = border
        
        # Datos
        for row_idx, obj in enumerate(queryset, 4):
            datos = [
                obj.id,
                f"{obj.valor} {unidad}",
                obj.fecha_hora.strftime("%Y-%m-%d %H:%M:%S"),
                obj.sensor_id
            ]
            
            for col_idx, valor in enumerate(datos, 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.value = valor
                cell.fill = row_fill if row_idx % 2 == 0 else PatternFill()
                cell.border = border
                cell.alignment = Alignment(horizontal='center')
        
        # Ajustar anchos
        ws.column_dimensions['A'].width = 8
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 20
        ws.column_dimensions['D'].width = 15